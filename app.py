import streamlit as st
import pandas as pd
import datetime
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage # Librería para insertar logos en Excel

# Configuración de la página web
st.set_page_config(page_title="Control de Asistencias - Sistema Corporativo", layout="wide")

if 'modo_impresion' not in st.session_state:
    st.session_state['modo_impresion'] = False

# Estilos visuales sobrios
st.markdown("""
    <style>
    html, body, [data-testid="stSidebar"] {
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    h1, h2, h3 {
        color: #4A3525 !important;
        font-weight: 600 !important;
    }
    [data-testid="stMetricValue"] {
        color: #D4AF37 !important;
        font-size: 24px !important;
    }
    [data-testid="stMetricLabel"] {
        color: #555555 !important;
        font-size: 14px !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .alerta-jornada {
        background-color: #FFF9C4 !important;
        color: #5D4037 !important;
        padding: 6px;
        border-radius: 4px;
        font-weight: bold;
        text-align: center;
    }
    @media print {
        @page { size: letter landscape; margin: 1cm; }
        [data-testid="stSidebar"], [data-testid="stFileUploader"], .no-print, header, footer, .stSelectbox, .stTextArea, button, .stDownloadButton, .element-container, .stButton {
            display: none !important;
        }
        .main .block-container { padding-top: 0rem !important; padding-bottom: 0rem !important; width: 100% !important; }
        table { width: 100% !important; page-break-inside: avoid; }
    }
    </style>
""", unsafe_allow_html=True)

if 'lista_horarios' not in st.session_state:
    st.session_state['lista_horarios'] = {
        "Descanso Obligatorio": {"entrada": datetime.time(0, 0), "salida": datetime.time(0, 0)}
    }

if 'asignacion_filas' not in st.session_state:
    st.session_state['asignacion_filas'] = {}

if 'observaciones_diarias' not in st.session_state:
    st.session_state['observaciones_diarias'] = {}

# Nombre del archivo de logo unificado
logo_file_name = 'image_963f8c.png'

with st.sidebar:
    if os.path.exists(logo_file_name):
        st.image(logo_file_name, use_container_width=True)
    else:
        st.markdown("<h2 style='text-align: center; color: #4A3525;'>SISTEMA</h2>", unsafe_allow_html=True)
    
    st.markdown("<hr style='border-top: 1px solid #ddd;'>", unsafe_allow_html=True)
    st.markdown("<b style='color: #4A3525;'>1. Cargar Archivo del Checador</b>", unsafe_allow_html=True)
    archivo_subido = st.file_uploader("Subir (.xls, .xlsx)", type=["xls", "xlsx"], label_visibility="collapsed")
    
    if st.session_state['modo_impresion']:
        if st.button("🔄 Volver a Vista de Edición", use_container_width=True):
            st.session_state['modo_impresion'] = False
            st.rerun()

def string_a_time(time_str):
    try:
        if pd.isna(time_str) or time_str == "" or time_str == "nan":
            return None
        parts = str(time_str).strip().split(':')
        return datetime.time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts)>2 else 0)
    except:
        return None

def diferencia_minutos(t1, t2):
    if not t1 or not t2: return 0
    return (t1.hour * 60 + t1.minute) - (t2.hour * 60 + t2.minute)

dias_espanol = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes', 5: 'Sábado', 6: 'Domingo'}

st.header("Reporte Ejecutivo de Asistencias e Incidencias")

if not st.session_state['modo_impresion']:
    st.markdown("### ⚙️ Definición del Catálogo de Horarios")
    with st.expander("Abrir Administrador de Horarios", expanded=True):
        c_crear, c_ver = st.columns([1, 1])
        with c_crear:
            nombre_del_turno = st.text_input("Nombre del Turno:", key="admin_nom_turno")
            col_t1, col_t2 = st.columns(2)
            with col_t1: h_entrada = st.time_input("Hora de Entrada:", datetime.time(9, 0), key="admin_ent_turno")
            with col_t2: h_salida = st.time_input("Hora de Salida:", datetime.time(18, 0), key="admin_sal_turno")
            if st.button("💾 Registrar Horario en el Sistema", use_container_width=True):
                if nombre_del_turno:
                    clave = f"{nombre_del_turno.strip()} ({h_entrada.strftime('%H:%M')} a {h_salida.strftime('%H:%M')})"
                    st.session_state['lista_horarios'][clave] = {"entrada": h_entrada, "salida": h_salida}
                    st.success("Turno guardado exitosamente.")
                    st.rerun()
        with c_ver:
            st.markdown("<b>Tus Horarios Guardados:</b>", unsafe_allow_html=True)
            for t_clave in list(st.session_state['lista_horarios'].keys()):
                col_t_nom, col_t_btn = st.columns([4, 1])
                with col_t_nom: st.write(f"• {t_clave}")
                with col_t_btn:
                    if t_clave != "Descanso Obligatorio" and st.button("🗑️", key=f"borrar_{t_clave}"):
                        del st.session_state['lista_horarios'][t_clave]
                        st.rerun()

st.markdown("---")

if archivo_subido is not None:
    try:
        if archivo_subido.name.endswith('.xls'):
            dfs = pd.read_html(archivo_subido)
            df = dfs[0]
        else:
            df = pd.read_excel(archivo_subido)
            
        df['Fechacompleta'] = df['Fechacompleta'].astype(str)
        df['Hora'] = df['Hora'].astype(str)

        grouped = df.groupby(['Empleado', 'Fechacompleta', 'Centrodecosto', 'Puesto']).agg(
            Horas_Registradas=('Hora', list)
        ).reset_index()

        def procesar_jornada(row):
            horas = sorted(row['Horas_Registradas'])
            entrada = horas[0]
            salida = horas[-1] if len(horas) > 1 else ""
            return pd.Series([entrada, salida])

        grouped[['Entrada_Real', 'Salida_Real']] = grouped.apply(procesar_jornada, axis=1)
        grouped = grouped.drop(columns=['Horas_Registradas'])

        if not st.session_state['modo_impresion']:
            c1, c2, c3 = st.columns(3)
            with c1: emp_filtro = st.selectbox("Colaborador:", ["-- Todos los Colaboradores --"] + sorted(list(grouped['Empleado'].unique())))
            with c2: suc_filtro = st.selectbox("Sucursal:", ["-- Todas las Sucursales --"] + sorted(list(grouped['Centrodecosto'].unique())))
            with c3: pto_filtro = st.selectbox("Puesto:", ["-- Todos los Puestos --"] + sorted(list(grouped['Puesto'].unique())))
        else:
            emp_filtro, suc_filtro, pto_filtro = "-- Todos los Colaboradores --", "-- Todas las Sucursales --", "-- Todos los Puestos --"

        df_filtrado = grouped.copy()
        if emp_filtro != "-- Todos los Colaboradores --": df_filtrado = df_filtrado[df_filtrado['Empleado'] == emp_filtro]
        if suc_filtro != "-- Todas las Sucursales --": df_filtrado = df_filtrado[df_filtrado['Centrodecosto'] == suc_filtro]
        if pto_filtro != "-- Todos los Puestos --": df_filtrado = df_filtrado[df_filtrado['Puesto'] == pto_filtro]

        st.markdown("<b style='color:#4A3525;'>Matriz de Control Diaria e Incidencias</b>", unsafe_allow_html=True)
        
        opciones_turnos = list(st.session_state['lista_horarios'].keys())
        registros_finales = []
        tot_retardos, tot_extra, tot_faltante = 0, 0, 0
        
        for idx, row in df_filtrado.iterrows():
            colaborador = row['Empleado']
            fecha_str = row['Fechacompleta']
            try:
                partes = fecha_str.split('/')
                fecha_obj = datetime.date(int(partes[2]), int(partes[1]), int(partes[0]))
                nombre_dia = dias_espanol[fecha_obj.weekday()]
            except:
                nombre_dia = "Lunes"

            e_real = string_a_time(row['Entrada_Real'])
            s_real = string_a_time(row['Salida_Real'])
            clave_fila = f"{colaborador}_{fecha_str}"
            
            idx_previo = 0
            if clave_fila in st.session_state['asignacion_filas'] and st.session_state['asignacion_filas'][clave_fila] in opciones_turnos:
                idx_previo = opciones_turnos.index(st.session_state['asignacion_filas'][clave_fila])
            
            if not st.session_state['modo_impresion']:
                col_sel, col_datos, col_obs = st.columns([1.5, 3.5, 1.5])
                with col_sel:
                    turno_seleccionado = st.selectbox(f"H_{clave_fila}", opciones_turnos, index=idx_previo, key=f"sel_{clave_fila}", label_visibility="collapsed")
                    st.session_state['asignacion_filas'][clave_fila] = turno_seleccionado
                with col_obs:
                    obs_previa = st.session_state['observaciones_diarias'].get(clave_fila, "")
                    nota_diaria = st.text_input(f"N_{clave_fila}", value=obs_previa, placeholder="Nota del día", key=f"txt_{clave_fila}", label_visibility="collapsed")
                    st.session_state['observaciones_diarias'][clave_fila] = nota_diaria
            else:
                turno_seleccionado = st.session_state['asignacion_filas'].get(clave_fila, opciones_turnos[0])
                nota_diaria = st.session_state['observaciones_diarias'].get(clave_fila, "")
                col_datos = st.container()

            conf = st.session_state['lista_horarios'].get(turno_seleccionado, {"entrada": datetime.time(0,0), "salida": datetime.time(0,0)})
            e_teorica = conf['entrada']
            s_teorica = conf['salida']
            
            mins_trabajados = diferencia_minutos(s_real, e_real) if (s_real and e_real) else 0
            if mins_trabajados < 0: mins_trabajados = 0
            
            style_alerta = "class='alerta-jornada'" if (turno_seleccionado != "Descanso Obligatorio" and mins_trabajados < 480) else ""
            
            retardo = 0
            if e_real and e_teorica and e_teorica != datetime.time(0,0):
                dif_ent = diferencia_minutos(e_real, e_teorica)
                if dif_ent >= 11:
                    retardo = dif_ent - 10
                    tot_retardos += retardo

            tiempo_extra, tiempo_faltante = 0, 0
            if e_teorica != datetime.time(0,0):
                if s_real and e_real:
                    if mins_trabajados > 480:
                        tiempo_extra = mins_trabajados - 480
                        tot_extra += tiempo_extra
                    elif mins_trabajados < 480:
                        tiempo_faltante = 480 - mins_trabajados
                        tot_faltante += tiempo_faltante
                elif not s_real and e_real:
                    tiempo_faltante = 480
                    tot_faltante += 480

            def txt_mins(m): return f"{m // 60}:{m % 60:02d}:00" if m > 0 else "N/A"
            texto_laborado = f"{mins_trabajados // 60}:{mins_trabajados % 60:02d}:00" if mins_trabajados > 0 else "00:00:00"

            with col_datos:
                st.markdown(f"""
                <table style='width:100%; border-collapse: collapse; margin-bottom: 5px; font-size: 13px;'>
                    <tr style='background-color: #FDFBF7; border-bottom: 1px solid #E6DFD3;'>
                        <td style='padding: 6px; width:22%;'><b>{colaborador}</b><br><small style='color:#777;'>{row['Centrodecosto']}</small></td>
                        <td style='padding: 6px; width:15%;'>{fecha_str} ({nombre_dia})</td>
                        <td style='padding: 6px; width:11%;'>Ent: {row['Entrada_Real'] if row['Entrada_Real'] else '--:--'}</td>
                        <td style='padding: 6px; width:11%;'>Sal: {row['Salida_Real'] if row['Salida_Real'] else 'Falta'}</td>
                        <td style='padding: 6px; width:11%;' {style_alerta}>{texto_laborado}</td>
                        <td style='padding: 6px; width:10%; color: {"#D32F2F" if retardo > 0 else "inherit"};'>Ret: {f"{retardo} min" if retardo > 0 else "-"}</td>
                        <td style='padding: 6px; width:10%; color: {"#388E3C" if tiempo_extra > 0 else "inherit"};'>Ext: {txt_mins(tiempo_extra)}</td>
                    </tr>
                    {"<tr style='background-color:#FDFBF7;'><td colspan='7' style='padding: 2px 6px 4px 6px; color:#5D4037; font-size:11px;'><b>Nota:</b> " + nota_diaria + "</td></tr>" if nota_diaria else ""}
                </table>
                """, unsafe_allow_html=True)

            registros_finales.append({
                "Fecha": fecha_str,
                "Hora": row['Entrada_Real'] if row['Entrada_Real'] else "N/A",
                "Tiporegistro": "Entrada",
                "Empleado": colaborador,
                "Centrodecosto": row['Centrodecosto'],
                "Puesto": row['Puesto'],
                "Minutos de retrado": f"{retardo} min" if retardo > 0 else "N/A",
                "Horas de trabajo": texto_laborado,
                "Tiempo extra": txt_mins(tiempo_extra) if tiempo_extra > 0 else "",
                "Observaciones": nota_diaria
            })
            
            if row['Salida_Real']:
                registros_finales.append({
                    "Fecha": "",
                    "Hora": row['Salida_Real'],
                    "Tiporegistro": "Salida",
                    "Empleado": "",
                    "Centrodecosto": "",
                    "Puesto": "",
                    "Minutos de retrado": "",
                    "Horas de trabajo": "",
                    "Tiempo extra": "",
                    "Observaciones": ""
                })

        st.markdown("---")
        st.markdown("### 📥 Panel de Descarga y Generación de Reportes")
        c_down_excel, c_down_print = st.columns([1, 2])
        
        with c_down_excel:
            wb = Workbook()
            ws = wb.active
            ws.title = "CEDIS"
            ws.views.sheetView[0].showGridLines = True
            
            # --- INSERCIÓN DEL LOGO CORPORATIVO ---
            if os.path.exists(logo_file_name):
                img = OpenpyxlImage(logo_file_name)
                # Redimensionamos proporcionalmente para que encaje estético en la parte superior (aprox 75x75 px)
                img.width = 75
                img.height = 75
                ws.add_image(img, 'A2')
            
            # Espaciado y títulos institucionales corridos a la columna B para no encimarse con el logo
            ws.append([])
            ws.append([])
            ws.append([])
            ws.cell(row=4, column=2, value="MALINALLI EL SABOR DE LOS DIOSES, NIEVES Y PRODUCTOS CONGELADOS S.A. DE C.V")
            ws.cell(row=5, column=2, value="CONTROL DE ASISTENCIA Y PUNTUALIDAD")
            ws.append([])
            
            ws["B4"].font = Font(name="Calibri", size=11, bold=True)
            ws["B5"].font = Font(name="Calibri", size=11, bold=True, underline="single")
            
            # Altura personalizada para filas de la cabecera para albergar cómodamente el logo
            ws.row_dimensions[2].height = 25
            ws.row_dimensions[3].height = 25
            
            # Columnas del formato original
            headers = ["Fecha", "Hora", "Tiporegistro", "Empleado", "Centrodecosto", "Puesto", "Minutos de retrado", "Horas de trabajo", "Tiempo extra", "Observaciones"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True)
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )
            
            for col_num in range(1, 11):
                cell = ws.cell(row=7, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            
            for reg in registros_finales:
                row_data = [reg[h] for h in headers]
                ws.append(row_data)
                curr_row = ws.max_row
                for col_num in range(1, 11):
                    c = ws.cell(row=curr_row, column=col_num)
                    c.font = Font(name="Calibri", size=11)
                    c.border = thin_border
                    if col_num in [1, 2, 3]:
                        c.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.download_button(
                label="📥 Descargar Excel con Formato Corporativo",
                data=excel_buffer,
                file_name=f"CONTROL_DE_ASISTENCIA_{datetime.date.today().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        with c_down_print:
            if not st.session_state['modo_impresion']:
                if st.button("🖨️ Activar Vista Limpia para Impresión / PDF", use_container_width=True):
                    st.session_state['modo_impresion'] = True
                    st.rerun()
            else:
                st.info("Vista de impresión activa. Presiona Ctrl + P para Guardar como PDF.")

    except Exception as e:
        st.error(f"Error procesando los datos: {e}")
else:
    st.markdown("<p style='color: #777; font-size: 14px;'>Por favor, carga el archivo maestro del checador en la barra lateral para procesar los cálculos.</p>", unsafe_allow_html=True)